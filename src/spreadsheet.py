"""
Planilha de controle (Excel) com openpyxl.

Gera data/controle_operacoes.xlsx com as abas:
 Universo_B3, Pares_Encontrados, Pares_Filtrados, Sinais, Operacoes,
 Resumo_Mensal, Imposto, Posicoes, Backtest, Configuracoes.

Também oferece funções para registrar sinais e operações em CSV
(data/sinais.csv e data/operacoes.csv), usados como base das abas.
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src import get_paths
from src.logger import get_logger

logger = get_logger()

# Cabeçalhos padrão de cada aba que não vem direto de um CSV.
SINAIS_HEADERS = [
    "data", "hora", "empresa_base", "ticker_on", "ticker_pn", "ratio_atual",
    "media", "z_score", "sinal", "decisao", "motivo", "ganho_estimado",
    "custo_estimado", "imposto_estimado",
]

OPERACOES_HEADERS = [
    "data", "hora", "par", "ativo_vendido", "quantidade_vendida", "preco_venda",
    "total_venda", "ativo_comprado", "quantidade_comprada", "preco_compra",
    "total_compra", "custos", "imposto_estimado", "ganho_acoes",
    "resultado_liquido", "tipo_operacao", "observacoes",
]


# ---------------------------------------------------------------------------
# Registro incremental em CSV (fonte das abas Sinais e Operacoes)
# ---------------------------------------------------------------------------
def registrar_sinal_csv(config, signal):
    """Acrescenta um sinal ao data/sinais.csv."""
    paths = get_paths(config)
    agora = datetime.now()
    linha = {
        "data": agora.strftime("%Y-%m-%d"),
        "hora": agora.strftime("%H:%M:%S"),
        "empresa_base": signal.get("empresa_base", ""),
        "ticker_on": signal.get("ticker_on", ""),
        "ticker_pn": signal.get("ticker_pn", ""),
        "ratio_atual": signal.get("ratio_atual", ""),
        "media": signal.get("media_ratio", ""),
        "z_score": signal.get("z_score", ""),
        "sinal": signal.get("sinal", ""),
        "decisao": signal.get("decisao", ""),
        "motivo": signal.get("motivo", ""),
        "ganho_estimado": signal.get("ganho_estimado", ""),
        "custo_estimado": signal.get("custo_estimado", ""),
        "imposto_estimado": signal.get("imposto_estimado", ""),
    }
    _append_csv(paths["sinais"], linha, SINAIS_HEADERS)


def registrar_operacao_csv(config, signal, simulation, tipo_operacao="swing",
                           observacoes="paper trading"):
    """Acrescenta uma operação simulada ao data/operacoes.csv."""
    paths = get_paths(config)
    agora = datetime.now()
    sim = simulation or {}
    linha = {
        "data": agora.strftime("%Y-%m-%d"),
        "hora": agora.strftime("%H:%M:%S"),
        "par": sim.get("par", f"{signal.get('ticker_on')}/{signal.get('ticker_pn')}"),
        "ativo_vendido": sim.get("ativo_vendido", ""),
        "quantidade_vendida": sim.get("quantidade_vendida", ""),
        "preco_venda": sim.get("preco_venda", ""),
        "total_venda": sim.get("total_vendido", ""),
        "ativo_comprado": sim.get("ativo_comprado", ""),
        "quantidade_comprada": sim.get("quantidade_comprada", ""),
        "preco_compra": sim.get("preco_compra", ""),
        "total_compra": sim.get("total_comprado", ""),
        "custos": round(sim.get("custo_operacional", 0) + sim.get("slippage_estimado", 0), 2),
        "imposto_estimado": sim.get("imposto_estimado", ""),
        "ganho_acoes": sim.get("ganho_liquido_acoes", ""),
        "resultado_liquido": sim.get("ganho_financeiro_estimado", ""),
        "tipo_operacao": tipo_operacao,
        "observacoes": observacoes,
    }
    _append_csv(paths["operacoes"], linha, OPERACOES_HEADERS)


def _append_csv(path, linha, headers):
    """Acrescenta uma linha (dict) a um CSV, criando cabeçalho se preciso."""
    try:
        df_novo = pd.DataFrame([linha], columns=headers)
        if os.path.exists(path):
            df_novo.to_csv(path, mode="a", header=False, index=False, encoding="utf-8")
        else:
            df_novo.to_csv(path, index=False, encoding="utf-8")
    except Exception as exc:
        logger.error("Falha ao registrar em %s: %s", path, exc)


# ---------------------------------------------------------------------------
# Construção do Excel
# ---------------------------------------------------------------------------
def _read_csv_safe(path):
    if path and os.path.exists(path):
        try:
            return pd.read_csv(path)
        except Exception as exc:
            logger.warning("Falha ao ler %s: %s", path, exc)
    return pd.DataFrame()


def _write_df_to_sheet(ws, df, headers=None):
    """Escreve um DataFrame em uma worksheet com cabeçalho estilizado."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")

    if df is None or df.empty:
        cols = headers if headers else (list(df.columns) if df is not None else [])
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=1, column=j, value=str(col))
            c.font = header_font
            c.fill = header_fill
        return

    cols = list(df.columns)
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=str(col))
        c.font = header_font
        c.fill = header_fill

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(cols, start=1):
            value = row[col]
            if pd.isna(value):
                value = ""
            ws.cell(row=i, column=j, value=value)

    # Largura automática simples.
    for j, col in enumerate(cols, start=1):
        largura = max(10, min(40, len(str(col)) + 2))
        ws.column_dimensions[get_column_letter(j)].width = largura


def _config_to_df(config):
    """Achata o config em um DataFrame chave/valor para a aba Configuracoes."""
    linhas = []
    for bloco, valores in config.items():
        if isinstance(valores, dict):
            for k, v in valores.items():
                linhas.append({"bloco": bloco, "parametro": k, "valor": str(v)})
        else:
            linhas.append({"bloco": bloco, "parametro": "", "valor": str(valores)})
    return pd.DataFrame(linhas)


def build_spreadsheet(config):
    """Gera o arquivo Excel completo. Devolve o caminho do arquivo."""
    paths = get_paths(config)

    wb = Workbook()
    # Remove a planilha padrão para criar as nossas em ordem.
    wb.remove(wb.active)

    abas = {
        "Universo_B3": _read_csv_safe(paths["universo"]),
        "Pares_Encontrados": _read_csv_safe(paths["pares"]),
        "Pares_Filtrados": _read_csv_safe(paths["pares_filtrados"]),
        "Sinais": _read_csv_safe(paths["sinais"]),
        "Operacoes": _read_csv_safe(paths["operacoes"]),
        "Resumo_Mensal": _resumo_mensal(paths["operacoes"]),
        "Imposto": _resumo_imposto(paths["operacoes"]),
        "Posicoes": pd.DataFrame(columns=[
            "data", "ativo", "quantidade", "preco_medio", "valor_atual",
            "observacoes"]),
        "Backtest": _read_csv_safe(paths["ranking"]),
        "Configuracoes": _config_to_df(config),
    }

    headers_padrao = {
        "Sinais": SINAIS_HEADERS,
        "Operacoes": OPERACOES_HEADERS,
    }

    for nome, df in abas.items():
        ws = wb.create_sheet(title=nome)
        _write_df_to_sheet(ws, df, headers=headers_padrao.get(nome))

    try:
        wb.save(paths["excel"])
        logger.info("Planilha Excel gerada em %s", paths["excel"])
    except PermissionError:
        logger.error("Não foi possível salvar o Excel (arquivo aberto?). "
                     "Feche %s e tente de novo.", paths["excel"])
    except Exception as exc:
        logger.error("Falha ao salvar Excel: %s", exc)

    return paths["excel"]


def _resumo_mensal(operacoes_path):
    """Resumo de resultado por mês a partir das operações."""
    df = _read_csv_safe(operacoes_path)
    if df.empty or "data" not in df.columns:
        return pd.DataFrame(columns=["mes", "operacoes", "resultado_liquido",
                                     "custos", "imposto_estimado"])
    df["mes"] = df["data"].astype(str).str.slice(0, 7)
    for col in ("resultado_liquido", "custos", "imposto_estimado"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            df[col] = 0
    agg = df.groupby("mes").agg(
        operacoes=("par", "count"),
        resultado_liquido=("resultado_liquido", "sum"),
        custos=("custos", "sum"),
        imposto_estimado=("imposto_estimado", "sum"),
    ).reset_index()
    return agg


def _resumo_imposto(operacoes_path):
    """Resumo de imposto por mês."""
    df = _read_csv_safe(operacoes_path)
    if df.empty or "data" not in df.columns:
        return pd.DataFrame(columns=["mes", "tipo_operacao", "imposto_estimado"])
    df["mes"] = df["data"].astype(str).str.slice(0, 7)
    if "imposto_estimado" in df.columns:
        df["imposto_estimado"] = pd.to_numeric(
            df["imposto_estimado"], errors="coerce").fillna(0)
    else:
        df["imposto_estimado"] = 0
    tipo = df["tipo_operacao"] if "tipo_operacao" in df.columns else "swing"
    df["tipo_operacao"] = tipo
    return df.groupby(["mes", "tipo_operacao"]).agg(
        imposto_estimado=("imposto_estimado", "sum")).reset_index()


if __name__ == "__main__":
    from src import load_config

    build_spreadsheet(load_config())
